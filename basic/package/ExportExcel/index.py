import os
import json
import math
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.pagebreak import Break, RowBreak, ColBreak
from openpyxl.styles import colors
from openpyxl.styles import Font, Color, PatternFill, Border, Side
from openpyxl.chart import (
    LineChart,
    BarChart,
    Reference,
    Series,
    series_factory
)

class ExportExcel:

    # 构造函数
    def __init__(self):
        pass
    
    # 初始化
    def install(self, *agrs):
        print(agrs)

    # 导出用车明细-会议编号
    def exportSheetByMice(self, dataList = []):
        # 记录当前工作簿有多少的sheet
        self.sheetnames = []
        try:
            # 尝试读取本地文件《会议统计分析.xlsx》
            wb = load_workbook(filename="会议统计分析.xlsx")
            # 每次运行给新sheet添加索引《sheet{index}》
            sheetnames = wb.sheetnames
            ws = wb.create_sheet("sheet{}".format(len(sheetnames)))
        except:
            # 读取本地文件《会议统计分析.xlsx》失败，就新建一个工作簿
            wb = Workbook()
            sheetnames = []
            ws = wb.active
        
        total = len(dataList) # x轴类别长度
        multiple = total/5 # 标准类别长度是5，所以total/5是标准宽度的倍数
        for row in dataList:
            ws.append(row) # 遍历list，向sheet添加数据

        # 金额
        c1 = BarChart()
        print(c1.width)
        c1.width = multiple*10+10
        c1.height = c1.height*1.5
        # x轴类别category，min_col=1,max_col=1 == 取自第一列（这种情况max_col可以省略不写）；min_row=2,max_row=6 意思是从前固定好的第一列，从第2行到第6行取数据作为category
        cats = Reference(ws, min_col=1, max_col=1, min_row=2, max_row=total)
        # 同理第一个图表取自第三列，第2行到第6行的数据作为该图表的数据
        v1 = Reference(ws, min_col=4, min_row=2, max_row=total)
        c1.add_data(v1, titles_from_data=False, from_rows=False)
        c1.set_categories(cats)
        c1.y_axis.axId = 100
        c1.x_axis.title = '会议编号'
        c1.y_axis.title = '金额/公里数'
        c1.y_axis.majorGridlines = None
        c1.title = '会议统计分析'
        # 自定义系列
        # c1.series = [Series(v1, title="用车花费")]


        # 订单量
        c2 = LineChart()
        v2 = Reference(ws, min_col=2, min_row=2, max_row=total)
        c2.add_data(v2, titles_from_data=False, from_rows=False) #
        c2.y_axis.axId = 200 # 可能多个公用一个x轴的y轴，所以为了区别不同的y轴，需要设置id，当然你也可以将不同的y轴的id设置一样，那么它们将公用一个y轴
        c2.y_axis.title = "订单量" # 该轴线的title
        c2.y_axis.crosses = "max" # 设置为max，那么该条轴线将位于系列右侧，默认左侧
        # c2.series = [Series(v2, title="使用次数")] # 自定义系列
        c1 += c2

        # 公里数
        c3 = LineChart()
        v3 = Reference(ws, min_col=3, min_row=2, max_row=total)
        c3.add_data(v3, titles_from_data=False, from_rows=False) #
        c3.y_axis.axId = 100 # 可能多个公用一个x轴的y轴，所以为了区别不同的y轴，需要设置id，当然你也可以将不同的y轴的id设置一样，那么它们将公用一个y轴
        c3.y_axis.title = "公里数" # 该轴线的title
        # c3.y_axis.crosses = "max" # 设置为max，那么该条轴线将位于系列右侧，默认左侧
        # c3.series = [Series(v3, title="使用次数")] # 自定义系列
        c1 += c3

        ws.add_chart(c1, "G2")

        # 排序
        ws.auto_filter.ref = "A1:D{}".format(total)
        ws.auto_filter.add_sort_condition("B2:B{}".format(total))


        wb.save("会议统计分析.xlsx")
        wb.close()
        print("文件生成成功")

    # 导出用车明细-供应商
    def exportSheetBySupplier(self, dataList = []):
        # 记录当前工作簿有多少的sheet
        self.sheetnames = []
        try:
            # 尝试读取本地文件《供应商服务分析.xlsx》
            wb = load_workbook(filename="供应商服务分析.xlsx")
            # 每次运行给新sheet添加索引《sheet{index}》
            sheetnames = wb.sheetnames
            ws = wb.create_sheet("sheet{}".format(len(sheetnames)))
        except:
            # 读取本地文件《供应商服务分析.xlsx》失败，就新建一个工作簿
            wb = Workbook()
            sheetnames = []
            ws = wb.active
        
        total = len(dataList) # x轴类别长度
        multiple = total/5 # 标准类别长度是5，所以total/5是标准宽度的倍数
        for row in dataList:
            ws.append(row) # 遍历list，向sheet添加数据
        # 合计
        lastRow = [None for x in dataList[0]]
        lastRow[0] = '总计'
        ws.append(lastRow)
        ws['C{}'.format(total+1)] = "=SUM(C2:C{})".format(total)
        ws['D{}'.format(total+1)] = "=SUM(D2:D{})".format(total)
        ws['E{}'.format(total+1)] = "=SUM(E2:E{})".format(total)
        ws['F{}'.format(total+1)] = "=SUM(F2:F{})".format(total)
        ws['G{}'.format(total+1)] = "=SUM(G2:G{})".format(total)
        # 合计样式
        lastRows = ws['A{}'.format(total+1):'G{}'.format(total+1)]
        def addFont(cell):
            # https://openpyxl.readthedocs.io/en/stable/styles.html
            # 字体样式
            cell.font = Font(color=colors.RED, italic=True, bold=True)
            # 填充样式
            cell.fill = PatternFill("solid", fgColor=colors.YELLOW)
        def addBorder(cell):
            # 边框样式
            border = Border(left=Side(border_style="thin",color="000000"),
                            right=Side(border_style="thin",color="000000"),
                            top=Side(border_style="thin",color="000000"),
                            bottom=Side(border_style="thin",color="000000"))
            cell.border = border
        [[addFont(cell) for cell in row]for row in lastRows]
        [[addBorder(cell) for cell in row] for row in ws['A1':'G{}'.format(total+1)]]


        # 结算金额
        c1 = BarChart()
        # c1.width = multiple*10
        cats = Reference(ws, min_col=1, max_col=1, min_row=2, max_row=total)
        v1 = Reference(ws, min_col=4, min_row=2, max_row=total)
        c1.add_data(v1, titles_from_data=False, from_rows=False)
        c1.set_categories(cats)
        c1.y_axis.axId = 100
        c1.x_axis.title = '供应商名称'
        c1.y_axis.title = '结算金额'
        c1.y_axis.majorGridlines = None
        c1.title = '供应商服务分析'
        # 自定义系列
        # c1.series = [Series(v1, title="用车花费")]


        # 服务次数
        c2 = LineChart()
        v2 = Reference(ws, min_col=3, min_row=2, max_row=total)
        c2.add_data(v2, titles_from_data=False, from_rows=False) #
        c2.y_axis.axId = 200 # 可能多个公用一个x轴的y轴，所以为了区别不同的y轴，需要设置id，当然你也可以将不同的y轴的id设置一样，那么它们将公用一个y轴
        c2.y_axis.title = "服务次数" # 该轴线的title
        c2.y_axis.crosses = "max" # 设置为max，那么该条轴线将位于系列右侧，默认左侧
        # c2.series = [Series(v2, title="使用次数")] # 自定义系列
        c1 += c2

        # 服务公里数
        c3 = LineChart()
        v3 = Reference(ws, min_col=5, min_row=2, max_row=total)
        c3.add_data(v3, titles_from_data=False, from_rows=False) #
        c3.y_axis.axId = 100 # 可能多个公用一个x轴的y轴，所以为了区别不同的y轴，需要设置id，当然你也可以将不同的y轴的id设置一样，那么它们将公用一个y轴
        c3.y_axis.title = "服务公里数" # 该轴线的title
        # c3.y_axis.crosses = "max" # 设置为max，那么该条轴线将位于系列右侧，默认左侧
        # c3.series = [Series(v3, title="使用次数")] # 自定义系列
        c1 += c3

        

        # 单公里价格
        c4 = BarChart()
        # c4.width = multiple*10
        cats = Reference(ws, min_col=1, max_col=1, min_row=2, max_row=total)
        v4 = Reference(ws, min_col=6, min_row=2, max_row=total)
        c4.add_data(v4, titles_from_data=False, from_rows=False)
        c4.set_categories(cats)
        c4.y_axis.axId = 100
        c4.x_axis.title = '供应商名称'
        c4.y_axis.title = '单公里价格'
        c4.y_axis.majorGridlines = None
        c4.title = '供应商单价对比'
        # 自定义系列
        # c4.series = [Series(v4, title="单公里价格")]

        # 均单价
        c5 = LineChart()
        # c5.width = multiple*10
        v5 = Reference(ws, min_col=7, min_row=2, max_row=total)
        c5.add_data(v5, titles_from_data=False, from_rows=False)
        c5.y_axis.axId = 300
        c5.y_axis.title = "均单价"
        c5.y_axis.crosses = "max"
        # 自定义系列
        # c5.series = [Series(v5, title="单公里价格")]
        c4 += c5

        # 向workSheet添加两张chart
        ws.add_chart(c1, "I1")
        ws.add_chart(c4, "I16")

        # 排序
        ws.auto_filter.ref = "A1:G5"
        # ws.auto_filter.add_filter_column(0, ["Kiwi", "Apple", "Mango"])
        ws.auto_filter.add_sort_condition("C2:C5")


        wb.save("供应商服务分析.xlsx")
        wb.close()
        print("文件生成成功") 

    # 导出用车明细-叫车人
    def exportSheetByUser(self, dataList = []):
        # 记录当前工作簿有多少的sheet
        self.sheetnames = []
        try:
            # 尝试读取本地文件《供应商服务分析.xlsx》
            wb = load_workbook(filename="供应商服务分析.xlsx")
            # 每次运行给新sheet添加索引《sheet{index}》
            sheetnames = wb.sheetnames
            ws = wb.create_sheet("sheet{}".format(len(sheetnames)))
        except:
            # 读取本地文件《供应商服务分析.xlsx》失败，就新建一个工作簿
            wb = Workbook()
            sheetnames = []
            ws = wb.active
        
        total = len(dataList) # x轴类别长度
        multiple = total/5 # 标准类别长度是5，所以total/5是标准宽度的倍数
        for row in dataList:
            ws.append(row) # 遍历list，向sheet添加数据

        # 用车花费
        c1 = BarChart()
        print(c1.width)
        c1.width = multiple*10
        # x轴类别category，min_col=1,max_col=1 == 取自第一列（这种情况max_col可以省略不写）；min_row=2,max_row=6 意思是从前固定好的第一列，从第2行到第6行取数据作为category
        cats = Reference(ws, min_col=1, max_col=1, min_row=2, max_row=total)
        # 同理第一个图表取自第三列，第2行到第6行的数据作为该图表的数据
        v1 = Reference(ws, min_col=3, min_row=2, max_row=total)
        c1.add_data(v1, titles_from_data=False, from_rows=False)
        c1.set_categories(cats)
        c1.x_axis.title = '叫车人姓名'
        c1.y_axis.title = '用车花费'
        c1.y_axis.majorGridlines = None
        c1.title = '供应商服务分析'
        # 自定义系列
        # c1.series = [Series(v1, title="用车花费")]


        # 使用次数
        c2 = LineChart()
        v2 = Reference(ws, min_col=2, min_row=2, max_row=total)
        c2.add_data(v2, titles_from_data=False, from_rows=False) #
        c2.y_axis.axId = 200 # 可能多个公用一个x轴的y轴，所以为了区别不同的y轴，需要设置id，当然你也可以将不同的y轴的id设置一样，那么它们将公用一个y轴
        c2.y_axis.title = "使用次数" # 该轴线的title
        c2.y_axis.crosses = "max" # 设置为max，那么该条轴线将位于系列右侧，默认左侧
        # c2.series = [Series(v2, title="使用次数")] # 自定义系列
        c1 += c2

        ws.add_chart(c1, "G8")

        # 排序
        ws.auto_filter.ref = "A1:E6"
        # ws.auto_filter.add_filter_column(0, ["Kiwi", "Apple", "Mango"])
        ws.auto_filter.add_sort_condition("B2:B6")


        wb.save("供应商服务分析.xlsx")
        wb.close()
        print("文件生成成功")