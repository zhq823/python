import os
import json
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.pagebreak import Break, PageBreak, RowBreak, ColBreak
from openpyxl.chart import (
    LineChart,
    BarChart,
    Reference,
    Series,
)

sheetnames = []
try:
    # 尝试读取本地文件《销售用车次数、花费.xlsx》
    wb = load_workbook(filename="销售用车次数、花费.xlsx")
    # 每次运行给新sheet添加索引《sheet{index}》
    sheetnames = wb.sheetnames
    ws = wb.create_sheet("sheet{}".format(len(sheetnames)))
except:
    # 读取本地文件《销售用车次数、花费.xlsx》失败，就新建一个工作簿
    wb = Workbook()
    sheetnames = []
    ws = wb.active

# 测试数据
dataList = [
    ["类别"],
    ['用车花费'],
    ['用车次数']
]
with open("{}/carAmount.json".format(os.path.dirname(__file__)), encoding='utf-8') as f:
    data = json.load(f)
    for item in data:
        dataList[0].append(item["name"])
        dataList[1].append(item["amount"])
        dataList[2].append(item["qty"])
    # print("json源数据", data)
    # print(dataList)
f.close()

# 总数量
total = len(dataList[0])
# print(total)

for row in dataList:
    ws.append(row)

# 第一个柱状图
c1 = BarChart()



cats = Reference(ws, min_col=2, max_col=total, min_row=1, max_row=1)
v1 = Reference(ws, min_col=1, min_row=2, max_col=total)
c1.add_data(v1, titles_from_data=True, from_rows=True)
c1.set_categories(cats)

c1.x_axis.title = '姓名'
c1.y_axis.title = '用车花费'
c1.y_axis.majorGridlines = None
c1.title = '销售用车次数、花费'


# Create a second chart
# 创建第二个图表
c2 = LineChart()
v2 = Reference(ws, min_col=1, min_row=3, max_col=total)
c2.add_data(v2, titles_from_data=True, from_rows=True)
c2.y_axis.axId = 200
c2.y_axis.title = "用车次数"

# Display y-axis of the second chart on the right by setting it to cross the x-axis at its maximum
# 将第二个图表的y轴设置为最大与x轴相交，从而在右侧显示该图表的y轴
c2.y_axis.crosses = "max"
c1 += c2

ws.add_chart(c1, "D8")

# TODO

try:
    ws.page_breaks.append(RowBreak(Break(id=5)))
except:
    ws.row_breaks.append(RowBreak(Break(id=5)))

# TODO

wb.save("销售用车次数、花费.xlsx")
print("文件生成成功")
