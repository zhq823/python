import os
import json
import math
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.pagebreak import Break, PageBreak, RowBreak, ColBreak
from openpyxl.chart import (
    LineChart,
    BarChart,
    Reference,
    Series,
)

sheetnames = []
try:
    # 尝试读取本地文件《供应商服务分析.xlsx》
    wb = load_workbook(filename="供应商服务分析.xlsx")
    # 每次运行给新sheet添加索引《sheet{index}》
    sheetnames = wb.sheetnames
    ws = wb.create_sheet("sheet{}".format(len(sheetnames)))
except:
    # 读取本地文件《供应商服务分析.xlsx》失败，就新建一个工作簿
    wb = Workbook()
    sheetnames = []
    ws = wb.active

# 测试数据
dataList = [
    ["类别"],
    ['结算金额'],
    ['服务公里数'],
    ['服务次数'],
    ['单公里价格'],
    ['均单价']
]
with open("{}/supplier.json".format(os.path.dirname(__file__)), encoding='utf-8') as f:
    data = json.load(f)
    # [data.extend(data) for x in range(5)]
    for item in data:
        dataList[0].append(item["name"]) # 车供应商
        dataList[1].append(item["amount"]) # 结算金额
        dataList[2].append(item["kilo"]) # 服务公里数
        dataList[3].append(item["qty"]) # 服务次数
        dataList[4].append(item["kPrice"]) # 单公里价格
        dataList[5].append(item["price"]) # 均单价
    # print("json源数据", data)
    # print(dataList)
f.close()

# 总数量
total = len(dataList[0])
multiple = math.ceil(total/5)
# print(total)
# print(math.ceil(total/5))

for row in dataList:
    ws.append(row)

# 第一个表: 结算金额
c1 = BarChart()


# TODO
# 分页符适用于打印，并不是展示用的
print(ws.row_breaks)
ws.row_breaks.append(Break(id=2))
ws.col_breaks.append(Break(id=2))
print(ws.row_breaks)
# TODO


c1.width = multiple*15
cats = Reference(ws, min_col=2, max_col=total, min_row=1, max_row=1)
v1 = Reference(ws, min_col=1, min_row=2, max_col=total)
c1.add_data(v1, titles_from_data=True, from_rows=True)
c1.y_axis.axId = 100
c1.set_categories(cats)
c1.x_axis.title = '供应商'
c1.y_axis.title = '结算金额/服务公里数'
c1.y_axis.majorGridlines = None
c1.title = '供应商服务分析'


# 第二个表: 服务公里数
c2 = LineChart()
v2 = Reference(ws, min_col=1, min_row=3, max_col=total)
c2.add_data(v2, titles_from_data=True, from_rows=True)
c2.y_axis.axId = 100
c2.y_axis.title = "服务公里数"
# Display y-axis of the second chart on the right by setting it to cross the x-axis at its maximum
# 将第二个图表的y轴设置为最大与x轴相交，从而在右侧显示该图表的y轴
c2.y_axis.crosses = "max"
c1 += c2

# 第三个表: 服务次数
c3 = LineChart()
v3 = Reference(ws, min_col=1, min_row=4, max_col=total)
c3.add_data(v3, titles_from_data=True, from_rows=True)
c3.y_axis.axId = 200
c3.y_axis.title = "用车次数"
c3.y_axis.crosses = "max"
c1 += c3

ws.add_chart(c1, "H10")
print("生成成功")

wb.save("供应商服务分析.xlsx")