from openpyxl import Workbook, load_workbook
from openpyxl.chart import (
    LineChart,
    BarChart,
    Reference,
    Series,
)

sheetnames = []
try:
    # 尝试读取本地文件《组合图.xlsx》
    wb = load_workbook(filename="组合图.xlsx")
    # 每次运行给新sheet添加索引《sheet{index}》
    sheetnames = wb.sheetnames
    ws = wb.create_sheet("sheet{}".format(len(sheetnames)))
except:
    # 读取本地文件《组合图.xlsx》失败，就新建一个工作簿
    wb = Workbook()
    sheetnames = []
    ws = wb.active


rows = [
    ["姓名","a1","a2","a3","a4","a5","a6"],
    ['张三', 2, 3, 4, 5, 6, 7],
    ['李四', 10, 40, 50, 20, 10, 50],
    ['王五', 5, 16, 34, 18, 20, 45],
]

for row in rows:
    ws.append(row)

# 第一个柱状图
c1 = BarChart()
cats = Reference(ws, min_col=2, max_col=7, min_row=1, max_row=1)
v1 = Reference(ws, min_col=1, min_row=2, max_col=7)
c1.add_data(v1, titles_from_data=True, from_rows=True)
c1.set_categories(cats)

c1.x_axis.title = 'Days'
c1.y_axis.title = 'Aliens'
c1.y_axis.majorGridlines = None
c1.title = '组合图'


# Create a second chart
# 创建第二个图表
c2 = LineChart()
v2 = Reference(ws, min_col=1, min_row=3, max_col=7)
c2.add_data(v2, titles_from_data=True, from_rows=True)
c2.y_axis.axId = 200
c2.y_axis.title = "Humans"

# Display y-axis of the second chart on the right by setting it to cross the x-axis at its maximum
# 将第二个图表的y轴设置为最大与x轴相交，从而在右侧显示该图表的y轴
c1.y_axis.crosses = "max"
c1 += c2

# 第三个图表
c3 = LineChart()
v3 = Reference(ws, min_col=1, min_row=4, max_col=7)
c3.add_data(v3, titles_from_data=True, from_rows=True)
c3.y_axis.axId = 300
c3.y_axis.title = "用车次数"

c1.y_axis.crosses = "max"
c1 += c3


ws.add_chart(c1, "D4")

wb.save("组合图.xlsx")