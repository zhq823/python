import os
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference, PieChart, LineChart

sheetnames = []
try:
    # 尝试读取本地文件《sheet2.xlsx》
    wb = load_workbook(filename="sheet2.xlsx")
    # 每次运行给新sheet添加索引《sheet{index}》
    sheetnames = wb.sheetnames
    sheet = wb.create_sheet("sheet{}".format(len(sheetnames)))
except:
    # 读取本地文件《sheet2.xlsx》失败，就新建一个工作簿
    wb = Workbook()
    sheetnames = []
    sheet = wb.active

# openpyxl写入工作簿的数据格式
dataList = [
    ['Number', 'Batch 1', 'Batch 2', 'Batch 3'],
    [2, 40, 30, 50],
    [3, 40, 25, 50],
    [4, 50, 30, 50],
    [5, 30, 10, 50],
    [6, 25, 5, 50],
    [7, 50, 10, 50],
]
for row in dataList:
    # 写入
    sheet.append(row)
print('数据写入成功！')

# 新建一个柱状图chart对象
chart = BarChart()

cats = Reference(sheet, min_col=1, min_row=2, max_row=7)
data = Reference(sheet, min_row=1, max_row=7, min_col=1, max_col=3)
chart.add_data(data, titles_from_data=True)
chart.set_categories(cats)
sheet.add_chart(chart, 'A10')
wb.save(filename='sheet2.xlsx')
print('柱状图绘制成功！')
