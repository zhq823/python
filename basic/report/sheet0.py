import os
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference, PieChart, LineChart

sheetnames = []
try:
    # 尝试读取本地文件《sheet0.xlsx》
    wb = load_workbook(filename="sheet0.xlsx")
    # 每次运行给新sheet添加索引《sheet{index}》
    sheetnames = wb.sheetnames
    sheet = wb.create_sheet("sheet{}".format(len(sheetnames)))
except:
    # 读取本地文件《sheet0.xlsx》失败，就新建一个工作簿
    wb = Workbook()
    sheetnames = []
    sheet = wb.active

# openpyxl写入工作簿的数据格式
dataList = [['小明', 1], ['小华', 2], ['小强', 3]]
for row in dataList:
    # 写入
    sheet.append(row)
print('数据写入成功！')

# 新建一个柱状图chart对象
chart = BarChart()
data = Reference(worksheet=sheet, min_row=1, max_row=20, min_col=2, max_col=2)
categories = Reference(sheet, min_col=1, min_row=1, max_row=20)

chart.add_data(data, titles_from_data=False)
chart.set_categories(categories)
sheet.add_chart(chart, 'D2')
# 保存到工作簿【注意：理论上每次操作只有一个save就可以了，类似于你在操作excel修改+保存】
wb.save(filename='sheet0.xlsx')
print('柱状图绘制成功！')

# 新建一个饼状图chart对象
pie = PieChart()
labels = Reference(sheet, min_col=1, min_row=1, max_row=20)
data = Reference(sheet, min_col=2, min_row=1, max_row=20)

pie.add_data(data, titles_from_data=False)
pie.set_categories(labels)
sheet.add_chart(pie,'M2')
# 保存到工作簿【注意：理论上每次操作只有一个save就可以了，类似于你在操作excel修改+保存】
wb.save(filename='sheet0.xlsx')
print('\n饼图绘制成功！')
